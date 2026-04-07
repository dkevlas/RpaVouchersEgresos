from openpyxl import load_workbook
import json, os
from src.config.path_folders import PathFolder


def sincronizar_y_corregir_discrepancias():
    try:
        filename_json = os.path.join(PathFolder.FOLDER_PROCESO, 'lista.json')
        excels = [f for f in os.listdir(PathFolder.FOLDER_PROCESO) if f.lower().endswith('.xlsx')]
        
        if not excels or not os.path.exists(filename_json):
            return

        excel_path = os.path.join(PathFolder.FOLDER_PROCESO, excels[0])
        
        with open(filename_json, "r", encoding="utf-8") as f:
            data_json = json.load(f)
            
        wb = load_workbook(excel_path)
        ws = wb.active
        hubo_cambios = False

        # Mapeo del JSON para búsqueda rápida por Serie y Número
        mapa_json = {(str(item['Serie']), str(item['Nro'])): idx for idx, item in enumerate(data_json)}

        for row in range(2, ws.max_row + 1):
            serie_ex = str(ws[f"B{row}"].value or "").strip()
            nro_ex = str(ws[f"C{row}"].value or "").strip()
            # Columna I: Procesado
            proc_ex = str(ws[f"I{row}"].value or "").strip().lower() 

            if not serie_ex or not nro_ex:
                continue

            clave = (serie_ex, nro_ex)
            
            # ... dentro del bucle de sincronización ...
            if clave in mapa_json:
                idx = mapa_json[clave]
                proc_js = str(data_json[idx].get('procesado', 'no')).lower()

                # SOLO actuar si el JSON ya dice que se procesó pero el Excel no se enteró
                if proc_js == "si" and proc_ex in ['none', '', 'no']:
                    estado_final = "si"
                    msg_duda = "Procesado con observaciones: validación manual requerida debido a falta de metadatos"

                    # 1. Actualizar JSON (asegurar observación)
                    if not data_json[idx].get('observacion'):
                        data_json[idx]['observacion'] = msg_duda
                    
                    # 2. Actualizar Excel
                    ws[f"I{row}"] = estado_final
                    if not ws[f"M{row}"].value:
                        ws[f"M{row}"] = msg_duda
                    
                    hubo_cambios = True

        if hubo_cambios:
            # Guardar JSON
            with open(filename_json, "w", encoding="utf-8") as f:
                json.dump(data_json, f, indent=4, ensure_ascii=False)
            # Guardar Excel
            wb.save(excel_path)
            print("✅ Discrepancias corregidas: Excel y JSON sincronizados.")
            
    except Exception as e:
        print(f"❌ Error crítico en sincronización: {e}")
        

def actualizar_json_por_excel() -> bool:
    try:
        sincronizar_y_corregir_discrepancias()
        
        print("En busca de actualizacion de json desde excel")
        carpeta = PathFolder.FOLDER_PROCESO
        
        excel = next(f for f in os.listdir(carpeta) if f.lower().endswith(".xlsx"))
        json_file = next(f for f in os.listdir(carpeta) if f.endswith(".json"))

        excel_path = os.path.join(carpeta, excel)
        json_path = os.path.join(carpeta, json_file)

        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        wb = load_workbook(excel_path, data_only=True)
        sheet = next((s for s in ["Sheet1", "Hoja1", "Hoja 1"] if s in wb.sheetnames), wb.sheetnames[0])
        ws = wb[sheet]

        # claves actuales en Excel (Serie, Nro)
        excel_keys = set()
        filas_excel = {}

        for row in range(2, ws.max_row + 1):
            serie = ws[f"B{row}"].value
            nro = ws[f"C{row}"].value
            procesada_excel = ws[f"I{row}"].value

            if serie and nro:
                key = (str(serie), str(nro))
                excel_keys.add(key)
                filas_excel[key] = procesada_excel

        nueva_data = []

        for item in data:
            key = (str(item.get("Serie")), str(item.get("Nro")))

            # existe en excel
            if key in excel_keys:
                # solo actualizar si NO procesado y excel vacío desde I
                if item.get("procesado") == "no" and not filas_excel[key]:
                    nueva_data.append(item)
                else:
                    nueva_data.append(item)
            else:
                # no existe en excel
                if item.get("procesado") == "si":
                    nueva_data.append(item)
                # si es "no" → se elimina

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(nueva_data, f, ensure_ascii=False, indent=4)

        return True
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f'Error al actualizar el json: {e}')
