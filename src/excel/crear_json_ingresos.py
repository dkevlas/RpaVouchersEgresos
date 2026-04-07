from openpyxl import load_workbook
from collections import defaultdict
from src.egresos.rutas.config import ConfigRutas
from datetime import datetime, timedelta
from pathlib import Path
import re
import json
import time


ENCABEZADOS = {
    "fecha": "fecha",
    "tip doc": "tipo_doc",
    "nro. docum.": "nro_documento",
    "orden de": "orden_de",
    "registro": "registro",
    "abono s/.": "importe",
    "glosa": "glosa",
}


def _normalizar(valor):
    if valor is None:
        return ""
    texto = str(valor).strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto.lower()


def _buscar_encabezados(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        mapa = {}
        for j, celda in enumerate(row):
            clave = _normalizar(celda)
            if clave in ENCABEZADOS:
                mapa[ENCABEZADOS[clave]] = j
        if len(mapa) >= len(ENCABEZADOS):
            return i, mapa
    return None, None


def excel_a_lotes_json_egresos(ruta_excel, hoja=None):

    wb = load_workbook(ruta_excel, data_only=True)

    if hoja:
        ws = wb[hoja]
    else:
        ws = wb.active

    fila_encabezado, columnas = _buscar_encabezados(ws)

    if columnas is None:
        print("No se encontraron los encabezados en el Excel")
        return []

    lotes = defaultdict(list)

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):

        if i <= fila_encabezado:
            continue

        fecha = row[columnas["fecha"]]
        tipo_doc = row[columnas["tipo_doc"]]
        nro_doc = row[columnas["nro_documento"]]
        orden_de = row[columnas["orden_de"]]
        registro = row[columnas["registro"]]
        importe = row[columnas["importe"]]
        glosa = row[columnas["glosa"]]

        if fecha is None:
            continue

        if isinstance(fecha, datetime):
            fecha_dt = fecha
        elif isinstance(fecha, (int, float)):
            # Serial de Excel: días desde 1899-12-30
            fecha_dt = datetime(1899, 12, 30) + timedelta(days=int(fecha))
        else:
            fecha_str_raw = str(fecha).strip()
            fecha_dt = datetime.strptime(fecha_str_raw, "%d/%m/%Y")

        fecha_str = fecha_dt.strftime("%d/%m/%Y")
        fecha_lote = fecha_dt.strftime("%d%m%y")

        registro_dict = {
            "fecha": fecha_str,
            "fecha_lote": fecha_lote,
            "tipo_doc": tipo_doc,
            "nro_documento": nro_doc,
            "orden_de": orden_de,
            "registro": "" if registro is None else registro,
            "importe": float(importe),
            "glosa": glosa,
            "fila": i
        }

        lotes[fecha_str].append(registro_dict)

    resultado = []

    for fecha, filas in lotes.items():

        primero = filas[0]
        orden_haber = primero["orden_de"]

        # haber: suma todos los importes con el mismo orden_de que la primera fila
        importe_haber = round(sum(f["importe"] for f in filas if f["orden_de"] == orden_haber), 2)

        haber = {
            "fecha": primero["fecha"],
            "fecha_lote": primero["fecha_lote"],
            "tipo_doc": primero["tipo_doc"],
            "nro_documento": primero["nro_documento"],
            "orden_de": primero["orden_de"],
            "importe": importe_haber,
            "glosa": primero["glosa"],
            "fila": primero["fila"]
        }

        # agregar: solo personas con orden_de diferente al de haber, consolidadas
        agregar = []
        for fila in filas:
            if fila["orden_de"] == orden_haber:
                continue
            existente = next(
                (f for f in agregar if f["orden_de"] == fila["orden_de"]),
                None
            )
            if existente:
                existente["importe"] = round(existente["importe"] + fila["importe"], 2)
            else:
                agregar.append({**fila})

        bloque = {
            "procesado": "no",
            "inicio": "",
            "fin": "",
            "observacion": "",
            "haber": haber,
            "deber": filas,
            "agregar": agregar
        }

        resultado.append(bloque)

    return resultado


def crear_json_egresos(ruta_excel, hoja=None):
    lotes = excel_a_lotes_json_egresos(ruta_excel, hoja)
    if lotes:
        ruta_json = Path(ConfigRutas.FOLDER_PROCESO) / "egresos.json"
        with open(ruta_json, "w", encoding="utf-8") as f:
            json.dump(lotes, f, ensure_ascii=False, indent=4)

        print(f"JSON creado exitosamente en: {ruta_json}")
        
        time.sleep(1)
        return ruta_json
