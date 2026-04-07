from openpyxl import load_workbook


def excel_a_json(path_excel: str) -> list:
    wb = load_workbook(path_excel, data_only=True)

    posibles = ["Sheet1", "Hoja1", "Hoja 1"]
    sheet_name = next((s for s in posibles if s in wb.sheetnames), wb.sheetnames[0])
    ws = wb[sheet_name]

    cabeceras = [c.value for c in ws[1]]
    data = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        item = dict(zip(cabeceras, row))

        item.update({
            "posicion": idx,
            "procesada": "no",
            "duracion": 0,
            "inicio": "",
            "final": "",
            "observacion": ""
        })

        data.append(item)

    return data
