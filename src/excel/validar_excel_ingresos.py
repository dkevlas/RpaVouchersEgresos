from openpyxl import load_workbook
from typing import TypedDict

class Return(TypedDict):
    success: bool
    message: str

CABECERAS_ESPERADAS = [
    "Fecha",
    "Serie",
    "Nro",
    "Cliente",
    "Tipo Pago",
    "Total",
    "Monto",
    "Nro Operación"
]


def validar_excel(path_excel: str) -> Return:
    try:
        wb = load_workbook(path_excel, data_only=True)

        # 1. Localizar la hoja correcta (Sheet1, Hoja1 o la primera disponible)
        posibles = ["Sheet1", "Hoja1", "Hoja 1"]
        sheet = next((wb[n] for n in posibles if n in wb.sheetnames), wb[wb.sheetnames[0]])

        # 2. Leer cabeceras limpiando espacios invisibles (.strip())
        # Solo leemos hasta la cantidad esperada para evitar errores por columnas extra
        cabeceras_excel = [
            str(cell.value).strip() if cell.value is not None else "" 
            for cell in sheet[1][:len(CABECERAS_ESPERADAS)]
        ]

        # 3. Validación detallada columna por columna
        for i, esperado in enumerate(CABECERAS_ESPERADAS):
            # Verificamos si la columna existe en el Excel
            if i >= len(cabeceras_excel):
                return {
                    "success": False,
                    "message": f"Estructura incompleta: Falta la columna '{esperado}'"
                }
            
            actual = cabeceras_excel[i]
            
            # Comparación exacta
            if actual != esperado:
                return {
                    "success": False,
                    "message": (f"Columna incorrecta en posición {i+1}. "
                                f"Esperado: '{esperado}', Encontrado: '{actual}'")
                }

        # 4. Verificación de integridad (si el Excel está vacío después de la cabecera)
        if sheet.max_row < 2:
            return {
                "success": False,
                "message": "El archivo Excel no contiene datos para procesar (solo cabeceras)."
            }

        return {
            "success": True,
            "message": "Excel válido"
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"Error técnico al validar el archivo: {str(e)}"
        }

def validar_excel_(path_excel: str) -> Return:
    try:
        wb = load_workbook(path_excel, data_only=True)

        posibles = ["Sheet1", "Hoja1", "Hoja 1"]
        sheet = None
        for nombre in posibles:
            if nombre in wb.sheetnames:
                sheet = wb[nombre]
                break

        if sheet is None:
            sheet = wb[wb.sheetnames[0]]

        # Leer cabeceras (fila 1)
        cabeceras = [cell.value for cell in sheet[1]]

        if cabeceras != CABECERAS_ESPERADAS:
            for i, esperado in enumerate(CABECERAS_ESPERADAS):
                if i >= len(cabeceras):
                    return {
                        "success": False,
                        "message": f"Falta la columna '{esperado}'"
                    }
                if cabeceras[i] != esperado:
                    return {
                        "success": False,
                        "message": f"Columna incorrecta en posición {i+1}. "
                                   f"Esperado '{esperado}', encontrado '{cabeceras[i]}'"
                    }

            return {
                "success": False,
                "message": "Las cabeceras no coinciden con el formato esperado"
            }

        return {
            "success": True,
            "message": "Excel válido"
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"Error al leer el Excel: {str(e)}"
        }
