import json
import os
from datetime import datetime
from openpyxl import load_workbook
from src.egresos.rutas.config import ConfigRutas


# Columnas de escritura en el Excel (base 1)
COL_PROCESADO = 9   # I
COL_INICIO = 10     # J
COL_FIN = 11        # K
COL_DURACION = 12   # L
COL_OBSERVACION = 13  # M


class EscritorEgresos:

    def __init__(self):
        self._ruta_json = os.path.join(ConfigRutas.FOLDER_PROCESO, "egresos.json")
        self._ruta_excel = self._buscar_excel()

    def _buscar_excel(self):
        archivos = [f for f in os.listdir(ConfigRutas.FOLDER_PROCESO) if f.lower().endswith(".xlsx")]
        if archivos:
            return os.path.join(ConfigRutas.FOLDER_PROCESO, archivos[0])
        return None

    def _leer_json(self):
        with open(self._ruta_json, "r", encoding="utf-8") as f:
            return json.load(f)

    def _guardar_json(self, data):
        with open(self._ruta_json, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

    def _escribir_celda(self, fila, columna, valor):
        if not self._ruta_excel:
            return
        wb = load_workbook(self._ruta_excel)
        ws = wb.active
        ws.cell(row=fila, column=columna, value=valor)
        wb.save(self._ruta_excel)

    def _buscar_lote(self, data, lote):
        haber = lote.get("haber", {})
        tipo_doc = str(haber.get("tipo_doc"))
        nro_doc = str(haber.get("nro_documento"))
        for item in data:
            h = item.get("haber", {})
            if str(h.get("tipo_doc")) == tipo_doc and str(h.get("nro_documento")) == nro_doc:
                return item
        return None

    def escribir_inicio(self, lote):
        ahora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        fila = lote["haber"]["fila"]

        # Escribir en Excel
        self._escribir_celda(fila, COL_PROCESADO, "si")
        self._escribir_celda(fila, COL_INICIO, ahora)

        # Escribir en JSON
        data = self._leer_json()
        item = self._buscar_lote(data, lote)
        if item:
            item["procesado"] = "si"
            item["inicio"] = ahora
            self._guardar_json(data)

        # Actualizar lote en memoria
        lote["procesado"] = "si"
        lote["inicio"] = ahora

    def escribir_fin(self, lote, observacion=""):
        ahora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        fila = lote["haber"]["fila"]

        # Calcular duración
        duracion_str = ""
        inicio_str = lote.get("inicio", "")
        if inicio_str:
            try:
                inicio_dt = datetime.strptime(inicio_str, "%d/%m/%Y %H:%M:%S")
                fin_dt = datetime.strptime(ahora, "%d/%m/%Y %H:%M:%S")
                total_seg = int((fin_dt - inicio_dt).total_seconds())
                minutos = total_seg // 60
                segundos = total_seg % 60
                duracion_str = f"{minutos:02d}:{segundos:02d}"
            except Exception:
                duracion_str = ""

        # Escribir en Excel
        self._escribir_celda(fila, COL_FIN, ahora)
        if duracion_str:
            self._escribir_celda(fila, COL_DURACION, duracion_str)

        # Escribir en JSON
        data = self._leer_json()
        item = self._buscar_lote(data, lote)
        if item:
            item["fin"] = ahora
            item["duracion"] = duracion_str
            if observacion:
                item["observacion"] = observacion
            self._guardar_json(data)

        # Actualizar lote en memoria
        lote["fin"] = ahora
        lote["duracion"] = duracion_str
        if observacion:
            lote["observacion"] = observacion

    def escribir_observacion(self, lote, observacion):
        fila = lote["haber"]["fila"]

        # Escribir en Excel
        self._escribir_celda(fila, COL_OBSERVACION, observacion)

        # Escribir en JSON
        data = self._leer_json()
        item = self._buscar_lote(data, lote)
        if item:
            item["observacion"] = observacion
            self._guardar_json(data)

        # Actualizar en memoria
        lote["observacion"] = observacion
