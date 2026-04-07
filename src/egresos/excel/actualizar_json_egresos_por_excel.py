from src.egresos.rutas.config import ConfigRutas
import os
import json
from openpyxl import load_workbook


def actualizar_json_por_excel():
    try:

        print("En busca de actualización de json desde excel")

        carpeta = ConfigRutas.FOLDER_PROCESO

        excel = next(f for f in os.listdir(carpeta) if f.lower().endswith(".xlsx"))
        json_file = next(f for f in os.listdir(carpeta) if f.endswith(".json"))

        excel_path = os.path.join(carpeta, excel)
        json_path = os.path.join(carpeta, json_file)

        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        wb = load_workbook(excel_path, data_only=True)

        sheet = next((s for s in ["Sheet1", "Hoja1", "Hoja 1"] if s in wb.sheetnames), wb.sheetnames[0])
        ws = wb[sheet]

        for lote in data:

            # si ya está procesado no revisamos nada
            if lote.get("procesado") == "si":
                continue

            filas_lote = []

            filas_lote.append(lote["haber"]["fila"])

            for d in lote["deber"]:
                filas_lote.append(d["fila"])

            lote_tocado = False

            for fila in filas_lote:

                valor_excel = ws[f"I{fila}"].value

                if valor_excel:  # hay observación o estado
                    lote_tocado = True
                    break

            if lote_tocado:
                lote["procesado"] = "si"

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

        return True

    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f'Error al actualizar el json: {e}')
        return False
