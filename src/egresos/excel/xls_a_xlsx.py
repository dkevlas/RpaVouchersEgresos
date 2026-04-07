import win32com.client as win32
import os
import xlrd
from openpyxl import Workbook
from datetime import datetime, timedelta


def xls_a_xlsx(ruta_xls: str):
    try:
        if not ruta_xls.lower().endswith(".xls"):
            return ruta_xls

        if not os.path.exists(ruta_xls):
            print(f"Archivo no encontrado: {ruta_xls}")
            return ruta_xls

        print("Convirtiendo XLS -> XLSX con xlrd + openpyxl")

        ruta_xlsx = ruta_xls + "x"

        xls_book = xlrd.open_workbook(ruta_xls)
        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name in xls_book.sheet_names():
            xls_sheet = xls_book.sheet_by_name(sheet_name)
            ws = wb.create_sheet(title=sheet_name)

            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    cell = xls_sheet.cell(row, col)
                    valor = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        dt_tuple = xlrd.xldate_as_tuple(valor, xls_book.datemode)
                        valor = datetime(*dt_tuple)
                        celda = ws.cell(row=row + 1, column=col + 1, value=valor)
                        celda.number_format = 'DD/MM/YYYY'
                    else:
                        ws.cell(row=row + 1, column=col + 1, value=valor)

        wb.save(ruta_xlsx)

        os.remove(ruta_xls)
        print(f"Convertido: {ruta_xlsx}")
        return ruta_xlsx
    except Exception as e:
        print(f"Error al convertir {ruta_xls}: {e}")
        return ruta_xls



def xls_a_xlsx_con_formato(ruta_xls: str):
    try:
        print(f"RUTA DEBUG: {ruta_xls}")
        if not ruta_xls.lower().endswith(".xls"):
             print("HOLA QUE HACE")
             return ruta_xls

        print(f"Verificando si existe el archivo: {os.path.exists(ruta_xls)}")

        print("Convirtiendo XLS -> XLSX (manteniendo formato)")

        ruta_xlsx = ruta_xls + "x"

        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(ruta_xls)

        print("HOla 3")

        wb.SaveAs(ruta_xlsx, FileFormat=51)  # 51 = xlsx
        wb.Close()
        print("hola 4")
        excel.Quit()

        os.remove(ruta_xls)
        print("hola 4")
        return ruta_xlsx
    except Exception as e:
        print(f"Error al convertir {ruta_xls} a {ruta_xlsx}: {e}")
        return ruta_xls
